#!/usr/bin/env python3
"""Build or append GEO metadata Excel from meta.json + trans.json.

Usage:
  # New Excel from scratch (raw GSE numbers → new file)
  python build_excel.py --meta meta.json --trans trans.json --output result.xlsx

  # Append to existing spreadsheet
  python build_excel.py --meta meta.json --trans trans.json --append existing.xlsx
"""

import argparse, json, sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Shared styles ──
HDR_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HDR_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
CELL_FONT = Font(name="Microsoft YaHei", size=10)
LINK_FONT = Font(name="Microsoft YaHei", size=10, color="0563C1", underline="single")
WRAP_TOP = Alignment(vertical="top", wrap_text=True)
CENTER_TOP = Alignment(horizontal="center", vertical="top")
OK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

NEW_HEADERS = [
    "Accession", "GEO标题(中)", "实验类型(中)", "GEO摘要(中)",
    "样本类型", "样本数", "PubMed IDs", "发布日期", "数据下载"
]

APPEND_HEADERS = [
    "GEO标题(中)", "实验类型(中)", "GEO摘要(中)",
    "样本类型", "NCBI样本数", "PubMed IDs", "发布日期", "数据下载"
]

COL_WIDTHS_NEW = {1: 14, 2: 44, 3: 20, 4: 65, 5: 28, 6: 10, 7: 20, 8: 14, 9: 14}
COL_WIDTHS_APPEND = {1: 44, 2: 20, 3: 65, 4: 28, 5: 12, 6: 20, 7: 14, 8: 14}


def build_row(acc, m, t):
    """Build a data row tuple from meta and translation dicts."""
    st = m.get("sample_types", [])
    return [
        acc,
        t.get("title_cn", ""),
        t.get("gdsType_cn", ""),
        t.get("summary_cn", ""),
        "、".join(st) if st else "",
        m.get("n_samples", ""),
        ", ".join(m.get("pmids", [])),
        m.get("PDAT", ""),
    ]


def style_header(ws, headers, row=1):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN; c.border = THIN


def style_data_row(ws, row, ncols, link_col):
    for ci in range(1, ncols + 1):
        c = ws.cell(row=row, column=ci)
        c.font = CELL_FONT
        c.alignment = WRAP_TOP
        c.border = THIN
    ws.cell(row=row, column=1).fill = OK_FILL
    for ci in (1, 5, 6, 7, 8, link_col):
        ws.cell(row=row, column=ci).alignment = CENTER_TOP


# ── New Excel ──
def build_new(meta, trans, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "GEO元数据"
    style_header(ws, NEW_HEADERS)

    for ri, acc in enumerate(meta, 2):
        m = meta[acc]
        t = trans.get(acc, {})
        vals = build_row(acc, m, t)
        for ci, v in enumerate(vals, 1):
            ws.cell(row=ri, column=ci, value=v)
        # Hyperlink
        geo_url = m.get("geo_url", "")
        c = ws.cell(row=ri, column=9)
        c.value = f'=HYPERLINK("{geo_url}", "GEO页面")'
        c.font = LINK_FONT; c.alignment = CENTER_TOP; c.border = THIN
        style_data_row(ws, ri, 9, 9)

    for ci, w in COL_WIDTHS_NEW.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28
    for ri in range(2, len(meta) + 2):
        ws.row_dimensions[ri].height = 120
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    return len(meta)


# ── Append to existing ──
def build_append(meta, trans, excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    # Find Accession column
    acc_col = None
    for ci in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=ci).value == "Accession":
            acc_col = ci
            break
    if acc_col is None:
        print("ERROR: No 'Accession' column found in spreadsheet.", file=sys.stderr)
        sys.exit(1)

    # Build accession → row index map
    acc_map = {}
    for ri in range(2, ws.max_row + 1):
        val = ws.cell(row=ri, column=acc_col).value
        if val:
            acc_map[str(val).strip()] = ri

    # Find last used column
    last_col = ws.max_column
    start_col = last_col + 1

    # Write headers
    for ci, h in enumerate(APPEND_HEADERS, start_col):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN; c.border = THIN

    # Write data
    matched = 0
    for acc, m in meta.items():
        if acc not in acc_map:
            print(f"  Warning: {acc} not found in spreadsheet, skipping.")
            continue
        ri = acc_map[acc]
        t = trans.get(acc, {})
        vals = build_row(acc, m, t)
        for ci, v in enumerate(vals, 1):
            ws.cell(row=ri, column=start_col + ci - 1, value=v)
        # Hyperlink (last append column)
        geo_url = m.get("geo_url", "")
        link_ci = start_col + len(APPEND_HEADERS) - 1
        c = ws.cell(row=ri, column=link_ci)
        c.value = f'=HYPERLINK("{geo_url}", "GEO页面")'
        c.font = LINK_FONT; c.alignment = CENTER_TOP; c.border = THIN
        matched += 1

    # Column widths
    for ci, w in COL_WIDTHS_APPEND.items():
        ws.column_dimensions[get_column_letter(start_col + ci - 1)].width = w

    # Save as new file (never overwrite original)
    out = excel_path.rsplit(".", 1)[0] + "_annotated.xlsx"
    wb.save(out)
    return matched


# ── Main ──
def main():
    p = argparse.ArgumentParser(description="Build or append GEO metadata Excel")
    p.add_argument("--meta", required=True, help="meta.json from fetch_geo.py")
    p.add_argument("--trans", required=True, help="trans.json from academic-translate")
    p.add_argument("--output", help="Output .xlsx path (new Excel mode)")
    p.add_argument("--append", help="Existing .xlsx to append columns to")
    args = p.parse_args()

    with open(args.meta, encoding="utf-8") as f:
        meta = json.load(f)
    with open(args.trans, encoding="utf-8") as f:
        trans = json.load(f)

    if args.append:
        n = build_append(meta, trans, args.append)
        print(f"Appended: {n} rows -> {args.append.rsplit('.', 1)[0]}_annotated.xlsx")
    elif args.output:
        n = build_new(meta, trans, args.output)
        print(f"Done: {n} rows -> {args.output}")
    else:
        print("ERROR: --output or --append required.", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
